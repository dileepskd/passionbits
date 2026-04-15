#!/usr/bin/env python3
"""Generate Passionbits market-research xlsx from per-sheet markdown sidecars.

Reads `.agents/prospect-lists/market-research/*.md` (one GFM table per file),
validates the schema, and writes a multi-sheet xlsx workbook with formatting.

Usage:
    python3 .agents/generate_market_research_xlsx.py            # generate xlsx
    python3 .agents/generate_market_research_xlsx.py --dry-run  # print row counts, no write
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

ROOT = Path(__file__).resolve().parent.parent
SIDECAR_DIR = ROOT / ".agents" / "prospect-lists" / "market-research"
OUTPUT_XLSX = ROOT / ".agents" / "prospect-lists" / "passionbits-market-research.xlsx"
VERIFICATION_TXT = SIDECAR_DIR / "verification-report.txt"

COMPANY_SCHEMA = [
    "Company", "Category", "HQ", "Target Market", "Why-fit signal",
    "Meta Ad Library URL", "Website", "LinkedIn", "Headcount",
    "Ad Spend signal", "UGC signal", "Source URL", "Confidence",
    "Outreach hook", "As-of date",
]

HEATMAP_SCHEMA = [
    "Region", "Category", "Spend signal ($)", "Top brands (3-5)",
    "Meta Ad Library sample URL", "Data source", "Confidence",
    "Hook angle", "As-of date",
]

SHEETS = [
    # (filename, sheet_title, schema, short_description)
    ("01-india-agencies-us-clients.md",   "1. India Agencies→US Clients",    COMPANY_SCHEMA,  "India-HQ agencies serving US clients (UGC, performance, influencer)"),
    ("02-us-ugc-agencies.md",             "2. US UGC Agencies",              COMPANY_SCHEMA,  "US agencies heavy on Meta UGC (production, influencer, creative testing)"),
    ("03-us-brands-to-india.md",          "3. US Brands→India",              COMPANY_SCHEMA,  "US brands expanding into India"),
    ("04-india-brands-to-us.md",          "4. India Brands→US",              COMPANY_SCHEMA,  "India D2C brands expanding to US"),
    ("05-consumer-apps-ugc.md",           "5. Consumer Apps on UGC",         COMPANY_SCHEMA,  "Consumer/prosumer apps running heavy Instagram UGC ads"),
    ("06-content-farms.md",               "6. Content Farms",                COMPANY_SCHEMA,  "Operators running content-farm playbooks"),
    ("07-singapore-to-us.md",             "7. Singapore→US",                 COMPANY_SCHEMA,  "Singapore companies expanding to US (UGC heavy)"),
    ("08-indonesia-to-us.md",             "8. Indonesia→US",                 COMPANY_SCHEMA,  "Indonesian companies expanding to US"),
    ("09-us-to-indonesia.md",             "9. US→Indonesia",                 COMPANY_SCHEMA,  "US companies expanding to Indonesia (D2C, beauty heavy)"),
    ("10-region-category-spend.md",       "10. Region×Category Spend",       HEATMAP_SCHEMA,  "Regional UGC/Meta ad-spend heatmap"),
    ("11-us-meta-ugc-brands.md",          "11. US UGC+Meta Heavy Brands",    COMPANY_SCHEMA,  "US brands heavy on UGC + Meta ad spend"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CONF_FILLS = {
    "High":   PatternFill("solid", fgColor="C6EFCE"),
    "Medium": PatternFill("solid", fgColor="FFEB9C"),
    "Low":    PatternFill("solid", fgColor="FFC7CE"),
}
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")

MD_TABLE_ROW_RE = re.compile(r"^\|(.+)\|\s*$")
MD_SEPARATOR_RE = re.compile(r"^\|\s*:?-+:?\s*(\|\s*:?-+:?\s*)+\|\s*$")


def parse_markdown_table(md_text: str, expected_headers: list[str], path: Path) -> list[list[str]]:
    """Extract the first GFM table from md_text. Returns list of row cell-lists (data rows only)."""
    lines = md_text.splitlines()
    rows: list[list[str]] = []
    header_idx = None
    for i, line in enumerate(lines):
        if MD_TABLE_ROW_RE.match(line) and i + 1 < len(lines) and MD_SEPARATOR_RE.match(lines[i + 1]):
            header_idx = i
            break
    if header_idx is None:
        return []
    header_cells = [c.strip() for c in lines[header_idx].strip().strip("|").split("|")]
    if header_cells != expected_headers:
        raise ValueError(
            f"Schema mismatch in {path.name}\n"
            f"  expected: {expected_headers}\n"
            f"  got:      {header_cells}"
        )
    for line in lines[header_idx + 2:]:
        if not MD_TABLE_ROW_RE.match(line):
            break
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) != len(expected_headers):
            continue
        if all(not c for c in cells):
            continue
        rows.append(cells)
    return rows


def autosize_columns(ws, schema: list[str], rows: list[list[str]], cap: int = 60) -> None:
    for col_idx, header in enumerate(schema, start=1):
        max_len = len(header)
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            if val and len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, cap)


def write_sheet(ws, schema: list[str], rows: list[list[str]]) -> None:
    for col_idx, header in enumerate(schema, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 28

    url_cols = {i + 1 for i, h in enumerate(schema) if "URL" in h or h in {"Website", "LinkedIn"}}
    conf_col = schema.index("Confidence") + 1 if "Confidence" in schema else None

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(schema, start=1):
            val = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP
            if c_idx in url_cols and val.startswith("http"):
                cell.hyperlink = val
                cell.font = LINK_FONT
            if conf_col and c_idx == conf_col and val in CONF_FILLS:
                cell.fill = CONF_FILLS[val]

    ws.freeze_panes = "B2"
    autosize_columns(ws, schema, rows)


def add_readme_sheet(wb: Workbook, sheet_stats: list[tuple[str, int]]) -> None:
    ws = wb.create_sheet("0. README", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90

    def put(row, a, b, bold_a=False):
        ca = ws.cell(row=row, column=1, value=a)
        cb = ws.cell(row=row, column=2, value=b)
        if bold_a:
            ca.font = Font(bold=True)
        ca.alignment = Alignment(vertical="top")
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    put(1, "Passionbits — Market Research", "", bold_a=True)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    put(2, "Generated", date.today().isoformat())
    put(3, "Source", "Markdown sidecars under .agents/prospect-lists/market-research/")
    put(4, "Script",  ".agents/generate_market_research_xlsx.py")
    put(5, "Skill",   "skills/market-research/")
    put(6, "", "")
    put(7, "Confidence legend", "", bold_a=True)
    put(8, "High",   "2+ independent sources, one of which is Meta Ad Library or a named press/funding item.")
    put(9, "Medium", "1 strong source (Meta Ad Library, Crunchbase, LinkedIn job post, reputable press).")
    put(10,"Low",    "Inference from adjacent signals only. Treat as lead, not fact.")
    put(11, "", "")
    put(12, "Sheets", "", bold_a=True)
    r = 13
    for title, n in sheet_stats:
        put(r, title, f"{n} rows")
        r += 1
    put(r + 1, "", "")
    put(r + 2, "How to use this file", "", bold_a=True)
    put(r + 3, "1. Sort by Category or Target Market to find your next outbound batch.")
    put(r + 4, "2. Filter Confidence = High first. Use Medium as a second wave.")
    put(r + 5, "3. Click Meta Ad Library URL to verify the ad-spend signal before contacting.")
    put(r + 6, "4. Copy the 'Outreach hook' into the first line of your email/LinkedIn DM.")

    for row in ws.iter_rows(min_row=1, max_row=r + 6):
        for c in row:
            if c.column == 1 and c.value and c.row > 12 and c.row <= 12 + len(sheet_stats):
                pass
    ws.sheet_view.showGridLines = False


def generate(dry_run: bool = False) -> int:
    if not SIDECAR_DIR.exists():
        print(f"ERROR: sidecar dir missing: {SIDECAR_DIR}", file=sys.stderr)
        return 2

    all_rows: dict[str, list[list[str]]] = {}
    row_counts: list[tuple[str, int]] = []
    missing_source: dict[str, int] = defaultdict(int)
    conf_dist: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    dedupe_map: dict[str, list[str]] = defaultdict(list)
    problems: list[str] = []

    for fname, title, schema, _desc in SHEETS:
        path = SIDECAR_DIR / fname
        if not path.exists():
            problems.append(f"[MISSING] {fname}")
            all_rows[title] = []
            row_counts.append((title, 0))
            continue
        try:
            rows = parse_markdown_table(path.read_text(encoding="utf-8"), schema, path)
        except ValueError as e:
            problems.append(str(e))
            rows = []
        all_rows[title] = rows
        row_counts.append((title, len(rows)))

        if schema is COMPANY_SCHEMA:
            src_idx = schema.index("Source URL")
            conf_idx = schema.index("Confidence")
            website_idx = schema.index("Website")
            for row in rows:
                if not (row[src_idx] or "").startswith("http"):
                    missing_source[title] += 1
                conf = row[conf_idx] if len(row) > conf_idx else ""
                conf_dist[title][conf or "Unknown"] += 1
                key = (row[website_idx] or row[0] or "").strip().lower()
                if key:
                    dedupe_map[key].append(title)
        else:
            for row in rows:
                conf_idx = schema.index("Confidence")
                conf = row[conf_idx] if len(row) > conf_idx else ""
                conf_dist[title][conf or "Unknown"] += 1

    allowed_overlap = {"2. US UGC Agencies", "11. US UGC+Meta Heavy Brands"}
    dedupe_conflicts = {
        k: v for k, v in dedupe_map.items()
        if len(v) > 1 and not (set(v) <= allowed_overlap)
    }

    total_rows = sum(n for _, n in row_counts)

    report_lines = [
        f"Passionbits Market Research — Verification Report",
        f"Generated: {date.today().isoformat()}",
        f"Total rows across all sheets: {total_rows}",
        "",
        "Row counts:",
    ]
    for title, n in row_counts:
        report_lines.append(f"  {title}: {n}")
    report_lines.append("")
    report_lines.append("Confidence distribution:")
    for title, _ in row_counts:
        dist = conf_dist.get(title, {})
        report_lines.append(f"  {title}: {dict(dist)}")
    report_lines.append("")
    if missing_source:
        report_lines.append("Rows missing Source URL (FAIL if > 0):")
        for title, n in missing_source.items():
            report_lines.append(f"  {title}: {n}")
    else:
        report_lines.append("All company rows have a Source URL. ✓")
    report_lines.append("")
    if dedupe_conflicts:
        report_lines.append("Cross-sheet dedupe conflicts:")
        for k, sheets in dedupe_conflicts.items():
            report_lines.append(f"  {k}: {sheets}")
    else:
        report_lines.append("No cross-sheet dedupe conflicts. ✓")
    report_lines.append("")
    if problems:
        report_lines.append("Problems:")
        report_lines.extend(f"  {p}" for p in problems)
    else:
        report_lines.append("No parse problems. ✓")

    report = "\n".join(report_lines)
    print(report)

    if dry_run:
        return 0 if not problems and not missing_source else 1

    VERIFICATION_TXT.parent.mkdir(parents=True, exist_ok=True)
    VERIFICATION_TXT.write_text(report, encoding="utf-8")

    wb = Workbook()
    wb.remove(wb.active)

    for fname, title, schema, _desc in SHEETS:
        ws = wb.create_sheet(title[:31])
        write_sheet(ws, schema, all_rows[title])
        if schema is HEATMAP_SCHEMA:
            spend_col_letter = get_column_letter(schema.index("Spend signal ($)") + 1)
            data_bar = DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True)
            last_row = max(2, len(all_rows[title]) + 1)
            ws.conditional_formatting.add(f"{spend_col_letter}2:{spend_col_letter}{last_row}", data_bar)

    add_readme_sheet(wb, row_counts)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\nWrote {OUTPUT_XLSX}")
    print(f"Wrote {VERIFICATION_TXT}")
    return 0 if not problems else 1


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--dry-run", action="store_true", help="Validate and print row counts without writing xlsx")
    args = p.parse_args()
    return generate(dry_run=args.dry_run)


if __name__ == "__main__":
    sys.exit(main())
